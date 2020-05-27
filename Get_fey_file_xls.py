import openpyxl
from sys import argv
import os

from openpyxl import Workbook

file_name = str(argv[1]).split("\\")[-1]

dir_file = str(argv[1])[:int(str(argv[1].__len__())) - file_name.__len__()]
os.chdir(dir_file)

wb = Workbook()

wb_with_date = openpyxl.load_workbook(filename=file_name)
sheet = wb_with_date[wb_with_date.sheetnames[0]]

array_manager = []
array_manager_info = []

os.chdir('C:\\Users\\Boni\\PycharmProjects\\Test_new_age\\resultFile')

for i in range(2, 1048000):

    if sheet.cell(row=i, column=1).value is None:
        break

    if sheet.cell(row=i, column=1).value not in array_manager:
        array_manager.append(sheet.cell(row=i, column=1).value)
        wb.save(filename=sheet.cell(row=i, column=1).value + '.xlsx')
        xls_manager = openpyxl.load_workbook(filename=sheet.cell(row=i, column=1).value + '.xlsx')

        active_sheet = xls_manager.active
        active_sheet.title = "Klient base"
        active_sheet.cell(row=1, column=1, value=sheet['B1'].value)
        active_sheet.cell(row=1, column=2, value=sheet['C1'].value)

        xls_manager.save(sheet.cell(row=i, column=1).value + '.xlsx')

    array_manager_info.append(
        [sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value])

array_manager_info = sorted(array_manager_info)

index = 2
previous_manager = array_manager_info[0][0]

xls_manager = openpyxl.load_workbook(filename=array_manager_info[0][0]+'.xlsx')
ws = xls_manager.active

for line_info in array_manager_info:
    if previous_manager != line_info[0]:
        index = 2
        xls_manager.save(filename=previous_manager + '.xlsx')
        previous_manager = line_info[0]

        xls_manager = openpyxl.load_workbook(filename=line_info[0] + '.xlsx')
        ws = xls_manager.active

    ws.cell(row=index, column=1, value=line_info[1])
    ws.cell(row=index, column=2, value=line_info[2])
    index += 1

xls_manager.save(filename=array_manager_info[-1][0]+'.xlsx')
